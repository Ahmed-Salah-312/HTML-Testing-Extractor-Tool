import os
import sys
import openpyxl
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import PatternFill, Font, Alignment
from front import Ui_MainWindow

class Back_End_Class(QtWidgets.QWidget, Ui_MainWindow):
    update_progress = QtCore.pyqtSignal(int)
    update_status = QtCore.pyqtSignal(str, int)
    process_finished = QtCore.pyqtSignal()

    def __init__(self, MainWindow):
        super().__init__()
        self.setupUi(MainWindow)
        self.Run_Button.clicked.connect(self.Extracted_Data)
        self.Folders_Button.clicked.connect(self.Select_Folders)
        self.Files_Button.clicked.connect(self.Select_Files)
        self.Status_list.clear()
        self.progressBar.setRange(0, 100)
        self.progressBar.setValue(0)
        self.progressBar.hide()
        self.Brows_Button.hide()
        self.textBrowser.hide()
        self.Path_label.hide()

        self.file_paths = []
        self.folder_paths = []
        self.processed_files = set()
        self.thread = None
        self.is_running = False

        self.update_progress.connect(self.update_progress_bar)
        self.update_status.connect(self.append_status_and_progress)
        self.process_finished.connect(self.on_processing_finished)

    def Select_Folders(self):
        self.Brows_Button.show()
        self.textBrowser.show()
        self.Path_label.setText("Selected Folders")
        self.Path_label.show()
        try:
            self.Brows_Button.clicked.disconnect()
        except TypeError:
            pass
        self.Brows_Button.clicked.connect(self.Brows_Folders)

    def Select_Files(self):
        self.Brows_Button.show()
        self.textBrowser.show()
        self.Path_label.setText("Selected Files")
        self.Path_label.show()
        try:
            self.Brows_Button.clicked.disconnect()
        except TypeError:
            pass
        self.Brows_Button.clicked.connect(self.Brows_Files)

    def Brows_Files(self):
        self.textBrowser.clear()
        file_names, _ = QFileDialog.getOpenFileNames(self, 'Open Files', 'Select Files', 'HTML files (*.html)')
        if file_names:
            if set(file_names) == set(self.file_paths):
                reply = QMessageBox.question(self, 'Confirm',
                                             "You have selected the same files again. Do you want to run the extraction for these files?",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return

            already_processed = [file for file in file_names if file in self.processed_files]
            if already_processed:
                reply = QMessageBox.question(self, 'Confirm',
                                             f"The following files have been processed before: {', '.join([os.path.basename(file) for file in already_processed])}. Do you want to run the extraction for these files again?",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    file_names = [file for file in file_names if file not in self.processed_files]

            if any(not file_name.lower().endswith('.html') for file_name in file_names):
                QMessageBox.critical(self, "Error",
                                     "One or more selected files are not HTML files. Please choose files with .html extension.")
                return

            self.file_paths = file_names
            self.textBrowser.setText("\n".join(file_names))

    def Brows_Folders(self):
        self.textBrowser.clear()
        folder_name = QFileDialog.getExistingDirectory(self, 'Select Folder', '')
        if folder_name:
            if os.path.isfile(folder_name):
                QMessageBox.critical(self, "Error", "You must choose a folder, not a file.")
                return

            if folder_name in self.folder_paths:
                reply = QMessageBox.question(self, 'Confirm',
                                             "You have selected the same folder again. Do you want to run the extraction for this folder?",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return

            self.folder_paths.append(folder_name)
            self.textBrowser.setText("\n".join(self.folder_paths))

    def Extracted_Data(self):
        if self.is_running:
            return

        if not self.file_paths and not self.folder_paths:
            QMessageBox.warning(self, "Warning", "Please select at least one HTML file or folder before running the extraction.")
            return

        self.progressBar.show()
        self.progressBar.setValue(0)

        files_to_reprocess = [file for file in self.file_paths if file in self.processed_files]
        if files_to_reprocess:
            reply = QMessageBox.question(self, 'Confirm',
                                         f"The following files have been processed before: {', '.join([os.path.basename(file) for file in files_to_reprocess])}. Do you want to reprocess these files?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return

        for folder in self.folder_paths:
            for root, _, files in os.walk(folder):
                for file in files:
                    if file.lower().endswith('.html'):
                        self.file_paths.append(os.path.join(root, file))

        self.is_running = True
        self.thread = FileProcessingThread(self.file_paths, self.processed_files)
        self.thread.update_progress.connect(self.update_progress_bar)
        self.thread.update_status.connect(self.append_status_and_progress)
        self.thread.process_finished.connect(self.on_processing_finished)
        self.thread.start()

    def update_progress_bar(self, value):
        self.progressBar.setValue(value)

    def append_status_and_progress(self, status_message, progress_value):
        self.Status_list.append(status_message)
        self.progressBar.setValue(progress_value)

    def on_processing_finished(self):
        self.progressBar.hide()
        QMessageBox.information(self, "Success", "Data extraction has been completed.")
        self.processed_files.update(self.file_paths)
        self.is_running = False

class FileProcessingThread(QtCore.QThread):
    update_progress = QtCore.pyqtSignal(int)
    update_status = QtCore.pyqtSignal(str, int)
    process_finished = QtCore.pyqtSignal()

    def __init__(self, file_paths, processed_files, parent=None):
        super().__init__(parent)
        self.file_paths = file_paths
        self.processed_files = processed_files

    def run(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Extracted Data"

        headers = ["File Name", "Step Number", "Description", "Expected Result", "Step Result", "Overall Result"]
        sheet.append(headers)
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 80
        sheet.column_dimensions['C'].width = 90
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30

        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        blue_fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')

        header_font = Font(name='Calibri', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[1]:
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = header_alignment

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")

        total_files = len(self.file_paths)

        for index, path in enumerate(self.file_paths):
            file_name = os.path.basename(path)
            chrome_service = Service()
            html_file = webdriver.Chrome(service=chrome_service, options=chrome_options)
            html_file.get(f"file:///{path}")

            try:
                rows = WebDriverWait(html_file, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, "//table/tbody/tr")))

                overall_test_result = "Failed"
                start_row = sheet.max_row + 1
                file_start_row = start_row

                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) == 5:
                        row_data = [file_name, cells[0].text, cells[1].text, cells[2].text, cells[4].text, ""]
                        sheet.append(row_data)
                        file_name = ""  # Only write file name once at the start of each set of steps
                    elif len(cells) == 1 and "Test Result : PASSED" in cells[0].text:
                        overall_test_result = "Passed"

                # Write overall result in the same row as the file name
                #if we are in the first row of the excel sheet make the file name aligned with the overall results

                if(file_start_row==1):
                    sheet.cell(row=file_start_row , column=6, value=overall_test_result)
                    if overall_test_result == "Passed":
                        sheet.cell(row=file_start_row, column=6).fill = green_fill
                    else:
                        sheet.cell(row=file_start_row, column=6).fill = red_fill
                else:
                    sheet.cell(row=file_start_row +1, column=6, value=overall_test_result)
                    if overall_test_result == "Passed":
                        sheet.cell(row=file_start_row+1, column=6).fill = green_fill
                    else:
                        sheet.cell(row=file_start_row+1, column=6).fill = red_fill

                # Write file name and its state in the Status_list
                status_message = f"{os.path.basename(path)}: Overall Test Result: {overall_test_result}"
                progress_value = int((index + 1) / total_files * 100)
                self.update_status.emit(status_message, progress_value)

                # Add an empty row between files' data
                sheet.append([])

            except Exception as e:
                status_message = f"Failed to extract data from {file_name}: {e}"
                progress_value = int((index + 1) / total_files * 100)
                self.update_status.emit(status_message, progress_value)

            finally:
                html_file.quit()

        output_excel_path = r"C:\Users\ZBOOK G3\Desktop\BrightSkies Intern\Youssef's session\File Extractor\Extracted Data.xlsx"
        wb.save(output_excel_path)

        self.update_status.emit(f"Data has been written to {output_excel_path}", 100)
        self.process_finished.emit()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    MainWindow.setWindowTitle("HTML Extractor")
    ui = Back_End_Class(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
